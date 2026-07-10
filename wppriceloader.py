"""
WordPress/WooCommerce Price Loader with Custom Meta Fields
Loads pricing data and custom meta fields from Excel file into WooCommerce products

Updates:
- Product price (regular_price)
- original_sku meta field (from Part Number)
- replacement_avail meta field (yes if New Part Number exists, no otherwise)
- replacement_sku meta field (from New Part Number column)
- date_updated meta field (current date)

Required Excel Columns:
- Part Number: Original product SKU
- Retail Price: Product price

Optional Excel Columns:
- New Part Number: Replacement SKU (sets replacement_avail to 'yes' if present)
- Updated: Resume tracking column - rows marked "written" or "NOT_FOUND" are skipped

Features:
- Parallel batch processing for maximum speed (up to 4 batches simultaneously)
- Configurable batch size (default: 99 products per batch)
- Resume capability: marks rows as "written" after successful update
- Automatically saves progress to Excel file
- CSV logging of all updated post IDs
- Graceful shutdown with Ctrl+C - saves progress before exiting
- API connectivity monitoring - stops if connection fails to prevent false NOT_FOUND marks
"""
import os
import json
import logging
import requests
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
from woocommerce import API
from concurrent.futures import ThreadPoolExecutor, as_completed
import signal
import sys

# Global shutdown flag
shutdown_requested = False

def signal_handler(sig, frame):
    """Handle Ctrl+C gracefully"""
    global shutdown_requested
    if not shutdown_requested:
        shutdown_requested = True
        logger.warning("\n" + "="*80)
        logger.warning("SHUTDOWN REQUESTED - Finishing current batch and saving progress...")
        logger.warning("Press Ctrl+C again to force quit (may corrupt Excel file)")
        logger.warning("="*80 + "\n")
    else:
        logger.error("\nFORCE QUIT - Excel file may be corrupted!")
        sys.exit(1)

# Register signal handler
signal.signal(signal.SIGINT, signal_handler)

# Load site URL and credentials from config.py / keys.txt
_base_dir = Path(__file__).resolve().parent
import sys
sys.path.insert(0, str(_base_dir))
from config import WORDPRESS_URL

_keys_file = _base_dir / 'keys.txt'
CONSUMER_KEY = CONSUMER_SECRET = None
with open(_keys_file, 'r', encoding='utf-8') as _f:
    _lines = [l.strip() for l in _f if l.strip()]
for _i, _line in enumerate(_lines):
    if 'Consumer key' in _line and _i + 1 < len(_lines):
        CONSUMER_KEY = _lines[_i + 1]
    if 'Consumer secret' in _line and _i + 1 < len(_lines):
        CONSUMER_SECRET = _lines[_i + 1]
if not CONSUMER_KEY or not CONSUMER_SECRET:
    raise RuntimeError("Could not load WooCommerce credentials from keys.txt")

WP_URL = WORDPRESS_URL

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('wp_pricing_loader.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

logger.info(f"Connecting to: {WP_URL}")

# Initialize WooCommerce API
wcapi = API(
    url=WP_URL,
    consumer_key=CONSUMER_KEY,
    consumer_secret=CONSUMER_SECRET,
    version="wc/v3",
    timeout=30
)

# Custom endpoint base URL
_CUSTOM_BASE = WP_URL.rstrip('/') + '/wp-json/custom/v1/products-by-sku'

# Batch configuration
MAX_BATCH_PRODUCTS = 99  # WooCommerce batch API limit is 100, use 99 for maximum throughput
MAX_PARALLEL_BATCHES = 4  # Number of batches to process in parallel (optimal: 3-4 for fast servers)
MAX_CONSECUTIVE_ERRORS = 5  # Stop script if this many consecutive API errors occur


def test_api_connection():
    """Test if the API is reachable before processing"""
    try:
        logger.info("Testing API connection...")
        # Try to get a simple product list to verify connectivity
        resp = wcapi.get('products', params={'per_page': 1})
        if resp.status_code in [200, 201]:
            logger.info("API connection test successful")
            return True
        else:
            logger.error(f"API connection test failed: HTTP {resp.status_code}")
            return False
    except Exception as e:
        logger.error(f"API connection test failed: {e}")
        return False


def find_products_by_original_sku(sku):
    """
    Use the custom/v1 GET endpoint to find all post IDs with original_sku=sku.
    Returns tuple: (status, products)
        status: 'success' | 'error' | 'not_found'
        products: list of dicts {id, parent_id, type, wc_sku} or empty list
    """
    try:
        # Use query parameters for auth (endpoint checks these first)
        params = {
            'original_sku': sku,
            'consumer_key': CONSUMER_KEY,
            'consumer_secret': CONSUMER_SECRET
        }
        resp = requests.get(
            _CUSTOM_BASE,
            params=params,
            timeout=30
        )
        if resp.status_code != 200:
            logger.error(f"Custom endpoint GET error for {sku}: HTTP {resp.status_code} — {resp.text[:200]}")
            return ('error', [])
        
        data = resp.json()
        products = data.get('products', [])
        
        if not products:
            return ('not_found', [])
        
        return ('success', products)
        
    except requests.exceptions.Timeout:
        logger.error(f"Timeout calling custom endpoint for {sku}")
        return ('error', [])
    except requests.exceptions.ConnectionError:
        logger.error(f"Connection error calling custom endpoint for {sku}")
        return ('error', [])
    except Exception as e:
        logger.error(f"Error calling custom endpoint for {sku}: {e}")
        return ('error', [])


def batch_update_products(batch_updates, post_ids_file=None):
    """
    Update multiple products in a single WooCommerce batch request.
    
    Args:
        batch_updates: List of dicts with keys: id, parent_id, update_data
        post_ids_file: Optional file handle to write updated post IDs
    
    Returns: (success_count, failed_count)
    """
    if not batch_updates:
        return 0, 0
    
    # Separate simple products from variations
    simple_updates = []
    variation_updates = {}  # Grouped by parent_id
    
    for item in batch_updates:
        if item['parent_id'] and item['parent_id'] > 0:
            parent_id = item['parent_id']
            if parent_id not in variation_updates:
                variation_updates[parent_id] = []
            variation_updates[parent_id].append({
                'id': item['id'],
                **item['update_data']
            })
        else:
            simple_updates.append({
                'id': item['id'],
                **item['update_data']
            })
    
    success_count = 0
    failed_count = 0
    
    # Batch update simple products
    if simple_updates:
        try:
            simple_ids = [p['id'] for p in simple_updates]
            logger.info(f"  Updating {len(simple_updates)} simple products: {simple_ids}")
            batch_data = {'update': simple_updates}
            resp = wcapi.post('products/batch', batch_data)
            if resp.status_code == 200:
                result = resp.json()
                updated = result.get('update', [])
                success_count += len(updated)
                updated_ids = [p['id'] for p in updated]
                logger.info(f"  [OK] Successfully updated simple products: {updated_ids}")
                
                # Write to post IDs log file
                if post_ids_file:
                    for pid in updated_ids:
                        post_ids_file.write(f"{pid},simple,{datetime.now().isoformat()}\n")
            else:
                logger.error(f"  Batch update failed: {resp.status_code} {resp.text[:200]}")
                failed_count += len(simple_updates)
        except Exception as e:
            logger.error(f"  Batch update exception: {e}")
            failed_count += len(simple_updates)
    
    # Batch update variations (grouped by parent)
    for parent_id, variations in variation_updates.items():
        try:
            variation_ids = [v['id'] for v in variations]
            logger.info(f"  Updating {len(variations)} variations for product {parent_id}: {variation_ids}")
            batch_data = {'update': variations}
            resp = wcapi.post(f'products/{parent_id}/variations/batch', batch_data)
            if resp.status_code == 200:
                result = resp.json()
                updated = result.get('update', [])
                success_count += len(updated)
                updated_ids = [v['id'] for v in updated]
                logger.info(f"  [OK] Successfully updated variations for product {parent_id}: {updated_ids}")
                
                # Write to post IDs log file
                if post_ids_file:
                    for vid in updated_ids:
                        post_ids_file.write(f"{vid},variation,{parent_id},{datetime.now().isoformat()}\n")
            else:
                logger.error(f"  Variation batch failed for product {parent_id}: {resp.status_code}")
                failed_count += len(variations)
        except Exception as e:
            logger.error(f"  Variation batch exception for product {parent_id}: {e}")
            failed_count += len(variations)
    
    return success_count, failed_count


def process_batch_wrapper(batch_data):
    """
    Wrapper for parallel batch processing
    Returns: (batch_id, success_count, failed_count, row_indices)
    """
    batch_id, updates, row_indices, post_ids_file = batch_data
    logger.info(f"[Batch {batch_id}] Processing {len(updates)} products...")
    success, failed = batch_update_products(updates, post_ids_file)
    logger.info(f"[Batch {batch_id}] Complete - Success: {success}, Failed: {failed}")
    return batch_id, success, failed, row_indices


def mark_row_as_written(excel_path, row_index, updated_col_index):
    """Mark a row as 'written' in the Excel file"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        # openpyxl uses 1-based indexing, pandas uses 0-based
        # row_index is from pandas (0-based), so add 2 (1 for header, 1 for 0-based)
        ws.cell(row=row_index + 2, column=updated_col_index + 1, value='written')
        wb.save(excel_path)
        wb.close()
    except Exception as e:
        logger.error(f"Failed to mark row {row_index} as written: {e}")


def mark_row_as_none(excel_path, row_index, updated_col_index):
    """Mark a row as 'NOT_FOUND' in the Excel file when no products found"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        # openpyxl uses 1-based indexing, pandas uses 0-based
        # row_index is from pandas (0-based), so add 2 (1 for header, 1 for 0-based)
        ws.cell(row=row_index + 2, column=updated_col_index + 1, value='NOT_FOUND')
        wb.save(excel_path)
        wb.close()
    except Exception as e:
        logger.error(f"Failed to mark row {row_index} as NOT_FOUND: {e}")


def prepare_product_updates(sku, price, new_part_number=None):
    """
    Look up products by original_sku and prepare update data for batch processing.
    
    Args:
        sku: Original SKU (Part Number)
        price: Product price
        new_part_number: Replacement SKU if available
    
    Returns: Tuple (status, update_items)
        status: 'success' | 'error' | 'not_found'
        update_items: List of update items for batching
    """
    status, products = find_products_by_original_sku(sku)
    
    if status == 'error':
        return ('error', [])
    
    if status == 'not_found' or not products:
        return ('not_found', [])

    price_str = f"{float(price):.2f}"
    current_date = datetime.now().strftime('%Y-%m-%d')
    
    # Build meta_data array for custom fields
    meta_data = [
        {'key': 'original_sku', 'value': str(sku)},
        {'key': 'date_updated', 'value': current_date}
    ]
    
    # Handle replacement fields
    if new_part_number and str(new_part_number).strip() and str(new_part_number).lower() not in ['nan', 'none', '']:
        meta_data.append({'key': 'replacement_avail', 'value': 'yes'})
        meta_data.append({'key': 'replacement_sku', 'value': str(new_part_number).strip()})
    else:
        meta_data.append({'key': 'replacement_avail', 'value': 'no'})
        meta_data.append({'key': 'replacement_sku', 'value': ''})

    update_items = []
    for p in products:
        update_items.append({
            'id': p.get('id'),
            'parent_id': p.get('parent_id', 0),
            'update_data': {
                'regular_price': price_str,
                'meta_data': meta_data
            },
            'sku': sku,
            'price': price_str,
            'new_part_number': new_part_number
        })
    
    return ('success', update_items)


def process_excel_file(excel_path):
    """Process Excel file and update prices and meta fields in WooCommerce with batch processing"""
    if not os.path.exists(excel_path):
        logger.error(f"Excel file does not exist: {excel_path}")
        return
    
    try:
        # Read Excel file
        logger.info(f"Reading Excel file: {excel_path}")
        df = pd.read_excel(excel_path)
        
        # Check for required columns
        if 'Part Number' not in df.columns:
            logger.error(f"Excel file missing 'Part Number' column. Available columns: {df.columns.tolist()}")
            return
        
        if 'Retail Price' not in df.columns:
            logger.error(f"Excel file missing 'Retail Price' column. Available columns: {df.columns.tolist()}")
            return
        
        # Check for optional columns
        has_new_part_number = 'New Part Number' in df.columns
        has_updated_column = 'Updated' in df.columns
        
        if has_new_part_number:
            logger.info("Found 'New Part Number' column - will update replacement fields")
        else:
            logger.info("No 'New Part Number' column found - replacement fields will be set to 'no'")
        
        if has_updated_column:
            logger.info("Found 'Updated' column - will skip rows marked 'written' or 'NOT_FOUND' and track progress")
            updated_col_index = df.columns.get_loc('Updated')
        else:
            logger.info("No 'Updated' column found - adding it for progress tracking")
            df['Updated'] = ''
            updated_col_index = len(df.columns) - 1
            # Save the Excel file with the new column
            df.to_excel(excel_path, index=False)
        
        # Clean data
        df['Part Number'] = df['Part Number'].astype(str).str.strip()
        df = df.dropna(subset=['Part Number', 'Retail Price'])
        
        # Clean optional columns
        if has_new_part_number:
            df['New Part Number'] = df['New Part Number'].astype(str).str.strip()
            df['New Part Number'] = df['New Part Number'].replace('nan', '')
        
        # Filter out already processed rows (written or not_found)
        if has_updated_column:
            df['Updated'] = df['Updated'].astype(str).str.strip().str.lower()
            pending_df = df[~df['Updated'].isin(['written', 'not_found', 'none'])].copy()
            already_done = len(df) - len(pending_df)
            logger.info(f"Total rows: {len(df)}, Already processed: {already_done}, Pending: {len(pending_df)}")
        else:
            pending_df = df.copy()
            logger.info(f"Found {len(pending_df)} rows to process")
        
        if len(pending_df) == 0:
            logger.info("All rows already processed!")
            return
        
        # Test API connection before proceeding
        if not test_api_connection():
            logger.error("API connection test failed - cannot proceed with updates")
            logger.error("Please check your internet connection and API credentials")
            raise RuntimeError("API connection unavailable")
        
        # Open report file and post IDs log
        report_path = os.path.join(os.getcwd(), 'updated_prices.txt')
        post_ids_log_path = os.path.join(os.getcwd(), 'updated_post_ids.csv')
        
        with open(report_path, 'w', encoding='utf-8') as report_file, \
             open(post_ids_log_path, 'w', encoding='utf-8') as post_ids_file:
            
            # Write CSV header for post IDs log
            post_ids_file.write("post_id,type,parent_id,timestamp\n")
            
            report_file.write("WordPress Price Update Report (Batch Mode)\n")
            report_file.write("=" * 80 + "\n")
            report_file.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            report_file.write(f"Excel File: {excel_path}\n")
            report_file.write(f"Max Batch Size: {MAX_BATCH_PRODUCTS} products\n")
            report_file.write(f"Updates: Price + Custom Meta Fields (original_sku, replacement_avail, replacement_sku, date_updated)\n")
            report_file.write("=" * 80 + "\n\n")
            
            success_count = 0
            error_count = 0
            skipped_count = 0
            consecutive_errors = 0  # Track consecutive API errors
            
            batch_updates = []
            batch_row_indices = []
            pending_batches = []  # Queue of batches ready for parallel processing
            batch_counter = 0
            
            def submit_pending_batches():
                """Submit all pending batches for parallel processing"""
                nonlocal success_count, error_count, batch_counter
                
                if not pending_batches:
                    return
                
                logger.info(f"\n{'='*60}")
                logger.info(f"Submitting {len(pending_batches)} batches for parallel processing...")
                logger.info(f"{'='*60}")
                
                with ThreadPoolExecutor(max_workers=MAX_PARALLEL_BATCHES) as executor:
                    # Submit all batches
                    futures = {}
                    for batch_data in pending_batches:
                        future = executor.submit(process_batch_wrapper, batch_data)
                        futures[future] = batch_data
                    
                    # Collect results as they complete
                    for future in as_completed(futures):
                        try:
                            batch_id, success, failed, row_indices = future.result()
                            success_count += success
                            error_count += failed
                            
                            # Mark rows as written if successful
                            if success > 0:
                                for row_idx in row_indices:
                                    mark_row_as_written(excel_path, row_idx, updated_col_index)
                        except Exception as e:
                            logger.error(f"Batch processing error: {e}")
                            error_count += len(futures[future][1])  # Count all as errors
                
                logger.info(f"Parallel batch complete. Total success: {success_count}, Total failed: {error_count}\n")
                pending_batches.clear()
            
            # Process each row
            for index, row in pending_df.iterrows():
                # Check for graceful shutdown request
                if shutdown_requested:
                    logger.warning(f"Shutdown requested - stopping after current batch")
                    logger.warning(f"Processed {index} rows before shutdown")
                    break
                
                sku = row['Part Number']
                price = row['Retail Price']
                
                # Get optional fields
                new_part_number = row.get('New Part Number', '') if has_new_part_number else None
                
                # Clean up empty values
                if new_part_number and str(new_part_number).strip() in ['', 'nan', 'None']:
                    new_part_number = None
                
                # Validate price
                try:
                    price_float = float(price)
                    if price_float <= 0:
                        logger.warning(f"Invalid price {price} for SKU {sku}, skipping")
                        skipped_count += 1
                        continue
                except (ValueError, TypeError):
                    logger.warning(f"Non-numeric price '{price}' for SKU {sku}, skipping")
                    skipped_count += 1
                    continue
                
                logger.info(f"Preparing SKU {sku} for batch update")
                
                # Prepare updates for this SKU
                try:
                    status, updates = prepare_product_updates(sku, price, new_part_number)
                    
                    if status == 'error':
                        consecutive_errors += 1
                        logger.error(f"API error for SKU {sku} (consecutive errors: {consecutive_errors})")
                        
                        if consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                            logger.error(f"\n{'='*80}")
                            logger.error(f"CRITICAL: {consecutive_errors} consecutive API errors detected!")
                            logger.error(f"This indicates a network or API connectivity problem.")
                            logger.error(f"Stopping script to prevent marking valid SKUs as NOT_FOUND.")
                            logger.error(f"Please check your connection and try again.")
                            logger.error(f"{'='*80}\n")
                            raise RuntimeError(f"Too many consecutive API errors ({consecutive_errors})")
                        
                        error_count += 1
                        continue
                    
                    if status == 'not_found':
                        consecutive_errors = 0  # Reset on successful API call
                        logger.warning(f"No products found with original_sku: {sku} - marking as NOT_FOUND")
                        mark_row_as_none(excel_path, index, updated_col_index)
                        skipped_count += 1
                        continue
                    
                    # Reset consecutive error counter on success
                    consecutive_errors = 0
                    
                    if not updates:
                        logger.warning(f"No products found with original_sku: {sku} - marking as NOT_FOUND")
                        mark_row_as_none(excel_path, index, updated_col_index)
                        skipped_count += 1
                        continue
                    
                    # If a single SKU has more than 100 products, split into chunks and process immediately
                    if len(updates) > MAX_BATCH_PRODUCTS:
                        logger.warning(f"SKU {sku} has {len(updates)} products - splitting into chunks")
                        for i in range(0, len(updates), MAX_BATCH_PRODUCTS):
                            chunk = updates[i:i + MAX_BATCH_PRODUCTS]
                            logger.info(f"Processing chunk {i//MAX_BATCH_PRODUCTS + 1} with {len(chunk)} products...")
                            success, failed = batch_update_products(chunk, post_ids_file)
                            success_count += success
                            error_count += failed
                        
                        # Mark row as written
                        mark_row_as_written(excel_path, index, updated_col_index)
                        replacement_info = f" | Replacement: {new_part_number}" if new_part_number else ""
                        report_file.write(f"SKU: {sku} | Price: £{price}{replacement_info} | Products: {len(updates)} (split into chunks)\n")
                        continue
                    
                    # Check if adding these products would exceed the batch limit
                    if batch_updates and (len(batch_updates) + len(updates)) > MAX_BATCH_PRODUCTS:
                        # Batch is full - add to pending queue
                        batch_counter += 1
                        pending_batches.append((batch_counter, batch_updates, batch_row_indices, post_ids_file))
                        logger.info(f"Batch {batch_counter} ready with {len(batch_updates)} products")
                        
                        # Reset for next batch
                        batch_updates = []
                        batch_row_indices = []
                        
                        # If we have enough batches, submit them for parallel processing
                        if len(pending_batches) >= MAX_PARALLEL_BATCHES:
                            submit_pending_batches()
                    
                    # Add to current batch
                    batch_updates.extend(updates)
                    batch_row_indices.append(index)
                    
                    replacement_info = f" | Replacement: {new_part_number}" if new_part_number else ""
                    report_file.write(f"SKU: {sku} | Price: £{price}{replacement_info} | Products: {len(updates)}\n")
                    
                except Exception as e:
                    logger.error(f"Error preparing SKU {sku}: {e}")
                    error_count += 1
                    continue
            
            # Add any remaining batch to pending queue
            if batch_updates:
                batch_counter += 1
                pending_batches.append((batch_counter, batch_updates, batch_row_indices, post_ids_file))
                logger.info(f"Final batch {batch_counter} ready with {len(batch_updates)} products")
            
            # Process all remaining batches
            if pending_batches:
                submit_pending_batches()
            
            # Write summary
            report_file.write("\n" + "=" * 80 + "\n")
            if shutdown_requested:
                report_file.write(f"Processing Summary (INTERRUPTED):\n")
            else:
                report_file.write(f"Processing Summary:\n")
            report_file.write(f"  Total pending rows: {len(pending_df)}\n")
            report_file.write(f"  Successfully updated: {success_count}\n")
            report_file.write(f"  Skipped (no match/invalid): {skipped_count}\n")
            report_file.write(f"  Errors: {error_count}\n")
            if shutdown_requested:
                report_file.write(f"\n  NOTE: Processing was interrupted by user.\n")
                report_file.write(f"  Run the script again to continue from where it stopped.\n")
        
        logger.info(f"\n{'='*80}")
        logger.info(f"Post IDs log saved to: {post_ids_log_path}")
        if shutdown_requested:
            logger.warning(f"Processing INTERRUPTED by user - progress saved safely!")
            logger.warning(f"Run the script again to continue processing remaining rows")
        else:
            logger.info(f"Processing complete!")
        logger.info(f"  Pending rows processed: {len(pending_df)}")
        logger.info(f"  Successfully updated: {success_count}")
        logger.info(f"  Skipped: {skipped_count}")
        logger.info(f"  Errors: {error_count}")
        logger.info(f"Report saved to: {report_path}")
        logger.info(f"Progress saved to Excel file - rows marked as 'written'")
    
    except Exception as e:
        logger.error(f"Error reading Excel file: {str(e)}")
        raise


def main():
    """Main function to run the WordPress pricing data loader"""
    import sys
    
    # Check for command-line argument
    if len(sys.argv) != 2:
        print("Usage: python wppriceloader.py <excel_file>")
        print("Example: python wppriceloader.py PRCJUL25.xlsx")
        print("\nExpected Excel Columns:")
        print("  - Part Number (required) - Original SKU")
        print("  - Retail Price (required) - Product price")
        print("  - New Part Number (optional) - Replacement SKU")
        print("  - Updated (optional) - Progress tracking column")
        print("\nUpdates:")
        print("  - Product price (regular_price)")
        print("  - original_sku meta field")
        print("  - replacement_avail meta field (yes/no)")
        print("  - replacement_sku meta field (from New Part Number)")
        print("  - date_updated meta field (current date)")
        print("\nFeatures:")
        print(f"  - Parallel batch processing ({MAX_BATCH_PRODUCTS} products per batch)")
        print(f"  - Up to {MAX_PARALLEL_BATCHES} batches processed simultaneously for maximum speed")
        print("  - Resume capability (skips rows marked 'written' or 'NOT_FOUND' in Updated column)")
        print("  - Automatic progress saving to Excel file")
        print("  - Graceful shutdown with Ctrl+C (saves progress before exiting)")
        print(f"  - API connectivity monitoring (stops after {MAX_CONSECUTIVE_ERRORS} consecutive errors)")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    
    # If it's a relative path, make it absolute based on current working directory
    if not os.path.isabs(excel_path):
        excel_path = os.path.join(os.getcwd(), excel_path)
    
    logger.info(f"Starting WordPress pricing data loading from Excel: {excel_path}")
    
    try:
        process_excel_file(excel_path)
        if shutdown_requested:
            logger.warning("WordPress pricing data loading interrupted safely - progress saved")
            sys.exit(0)
        else:
            logger.info("WordPress pricing data loading completed successfully")
    except Exception as e:
        logger.error(f"Fatal error during processing: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
